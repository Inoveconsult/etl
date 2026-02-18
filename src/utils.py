import pandas as pd
import re
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


# from openpyxl.styles import Font

#Função para criar dataframe lendo excel
def criar_dataframe(origem, skiprows=0):
    origem = Path(origem)
    
    if not origem.exists():
        raise FileNotFoundError(f"Arquivo não encontrado!: {origem}")
    
    df = pd.read_excel(origem, skiprows=skiprows)
    
    return df


#Função para renomear meses
def renomear_meses(df):
    mapa_meses = {
        "JAN": "JANEIRO", "JANEIRO": "JANEIRO",
        "FEV": "FEVEREIRO", "FEVEREIRO": "FEVEREIRO",
        "MAR": "MARÇO", "MARÇO": "MARÇO",
        "ABR": "ABRIL", "ABRIL": "ABRIL",
        "MAI": "MAIO", "MAIO": "MAIO",
        "JUN": "JUNHO", "JUNHO": "JUNHO",
        "JUL": "JULHO", "JULHO": "JULHO",
        "AGO": "AGOSTO", "AGOSTO": "AGOSTO",
        "SET": "SETEMBRO", "SETEMBRO": "SETEMBRO",
        "OUT": "OUTUBRO", "OUTUBRO": "OUTUBRO",
        "NOV": "NOVEMBRO", "NOVEMBRO": "NOVEMBRO",
        "DEZ": "DEZEMBRO", "DEZEMBRO": "DEZEMBRO"
    }

    novas_colunas = {}

    for col in df.columns:
        # Remove números e separadores (/, -, _)
        texto_limpo = re.split(r"[\/\-_]", col)[0]
        texto_limpo = texto_limpo.strip().upper()

        if texto_limpo in mapa_meses:
            novas_colunas[col] = mapa_meses[texto_limpo]
    
    df.columns = df.columns.str.strip().str.upper()

    return df.rename(columns=novas_colunas)

def classificar_por_faixas_unificada(serie, faixas, padrao="Sem classificação"):
    """
    Classifica uma série numérica com base em múltiplos intervalos por conceito.

    Cada conceito pode ter um ou mais intervalos:
    Exemplo de faixas:

        FAIXAS = [
            ("Regular", [(None, 9), (71, None)]),
            ("Suficiente", [(10, 29)]),
            ("Bom", [(30, 49)]),
            ("Ótimo", [(50, 70)])
        ]

    Parâmetros:
    -----------
    serie : pd.Series
        Série numérica a ser classificada.
    faixas : list
        Lista de tuplas (nome_conceito, lista_intervalos), onde cada intervalo é (minimo, maximo).
        None pode ser usado para mínimo ou máximo infinito.
    padrao : str
        Valor a ser atribuído caso a nota não se encaixe em nenhum intervalo.

    Retorna:
    --------
    pd.Series
        Série com os conceitos atribuídos, mantendo o índice original.
    """
    # converte valores inválidos para NaN
    serie = pd.to_numeric(serie, errors="coerce")

    condicoes = []
    valores = []

    for nome, intervalos in faixas:
        if not isinstance(intervalos, (list, tuple)):
            raise ValueError(f"Intervalos inválidos para '{nome}', deve ser lista ou tupla.")

        cond_total = pd.Series(False, index=serie.index)

        for intervalo in intervalos:
            if not (isinstance(intervalo, (list, tuple)) and len(intervalo) == 2):
                raise ValueError(f"Intervalo inválido em '{nome}': {intervalo}")

            minimo, maximo = intervalo

            if minimo is None:
                cond = serie <= maximo
            elif maximo is None:
                cond = serie >= minimo
            else:
                cond = serie.between(minimo, maximo)

            cond_total |= cond  # combina múltiplos intervalos para o mesmo conceito

        condicoes.append(cond_total)
        valores.append(nome)

    resultado = np.select(condicoes, valores, default=padrao)

    return pd.Series(resultado, index=serie.index)


def classificar_por_faixas(serie, faixas, padrao="Sem classificação"):
    condicoes = []
    valores = []

    for nome, minimo, maximo in faixas:
        if minimo is None:
            cond = serie <= maximo
        elif maximo is None:
            cond = serie >= minimo
        else:
            cond = serie.between(minimo, maximo)

        condicoes.append(cond)
        valores.append(nome)

    return np.select(condicoes, valores, default=padrao)

def classificar_mais_acesso(serie, faixas, padrao="Sem classificação"):
    serie = pd.to_numeric(serie, errors="coerce")

    condicoes = []
    valores = []

    for nome, intervalos in faixas:   # agora espera 2 elementos
        cond_total = pd.Series(False, index=serie.index)

        for minimo, maximo in intervalos:
            if minimo is None:
                cond = serie <= maximo
            elif maximo is None:
                cond = serie >= minimo
            else:
                cond = serie.between(minimo, maximo)

            cond_total |= cond

        condicoes.append(cond_total)
        valores.append(nome)

    return np.select(condicoes, valores, default=padrao)

def normalizar_codigo(df, coluna):
    """
    Normaliza códigos numéricos (INE, CNES, IBGE):
    Remove .0 e preserva NaN.
    """
    df = df.copy()
    df[coluna] = df[coluna].astype("Int64").astype(str)
    return df

def remover_colunas_por_nome(df, colunas, limite=None):
    """
    Remove colunas por nome, com limite opcional.

    Parâmetros:
    - df: DataFrame pandas
    - colunas: lista de nomes de colunas
    - limite: quantidade máxima de colunas a remover
    """
    df.columns = df.columns.str.strip() # Remove espaços em branco
    
    colunas_existentes = [c for c in colunas if c in df.columns]

    if limite:
        colunas_existentes = colunas_existentes[:limite]

    return df.drop(columns=colunas_existentes, errors="ignore")


def converte_numero(df, coluna):
    """
    Converte coluna numérica com vírgula decimal para float.
    """
    df = df.copy()
    df[coluna] = pd.to_numeric(
        df[coluna]
        .astype(str)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    )
    return df


def renomear_coluna(df, coluna_origem, coluna_destino):
    """
    Renomeia uma coluna de forma dinâmica.
    """
    df = df.copy()
    if coluna_origem not in df.columns:
        raise ValueError(
            f"Coluna '{coluna_origem}' não encontrada. "
            f"Disponíveis: {list(df.columns)}"
        )
    return df.rename(columns={coluna_origem: coluna_destino})

def adicionar_mes(df_base, df_mes, coluna_mes, chave="INE"):
    """
    Adiciona ou atualiza uma coluna mensal no DataFrame base
    e garante conversão numérica.
    """
    df_base = df_base.copy().set_index(chave)
    df_mes = df_mes.copy().set_index(chave)

    if coluna_mes not in df_mes.columns:
        raise ValueError(f"Coluna '{coluna_mes}' não encontrada em df_mes")

    if coluna_mes not in df_base.columns:
        df_base[coluna_mes] = pd.NA

    # adiciona/atualiza valores
    df_base.update(df_mes[[coluna_mes]])

    df_base = df_base.reset_index()

    # 🔴 PASSO CRÍTICO: normalizar SEMPRE
    df_base = converte_numero(df_base, coluna_mes)

    return df_base

def calcular_nota_final(df, coluna_conceito, peso, nome_coluna_saida="coluna"):
    mapa_conceitos = {
        "Regular": 0.25,
        "Suficiente": 0.50,
        "Bom": 0.75,
        "Ótimo": 1.00
    }

    df = df.copy()

    conceitos_invalidos = set(df[coluna_conceito].dropna()) - set(mapa_conceitos)
    if conceitos_invalidos:
        raise ValueError(f"Conceitos inválidos encontrados: {conceitos_invalidos}")

    df[nome_coluna_saida] = df[coluna_conceito].map(mapa_conceitos) * peso

    return df

def ordenar_por_nota_e_quadrimestre(
    df,
    coluna_nota="coluna",
    coluna_quadrimestre="RESULTADO QUADRIMESTRE",
    ordem_decrescente=True
):
    """
    Ordena o DataFrame primeiro pela nota final e,
    em caso de empate, pelo resultado do quadrimestre.
    """

    df = df.copy()

    # garantir que as colunas são numéricas
    df[coluna_nota] = pd.to_numeric(df[coluna_nota], errors="coerce")
    df[coluna_quadrimestre] = pd.to_numeric(df[coluna_quadrimestre], errors="coerce")

    asc = not ordem_decrescente

    df = df.sort_values(
        by=[coluna_nota, coluna_quadrimestre],
        ascending=[asc, asc]
    )

    return df

def consolidar_indicadores(dfs, notas):
    """
    Consolida múltiplos DataFrames de indicadores em um único DataFrame.

    Cada DataFrame deve conter as colunas: 'INE', 'NOME DA EQUIPE' e a coluna de nota.

    Parâmetros:
    -----------
    dfs : list of pd.DataFrame
        Lista de DataFrames a serem consolidados.
    notas : list of str
        Lista com o nome da coluna de nota correspondente para cada DataFrame.
        O tamanho deve ser igual ao da lista dfs.

    Retorna:
    --------
    pd.DataFrame
        DataFrame consolidado contendo:
        - INE
        - NOME DA EQUIPE
        - colunas de notas padronizadas (uma para cada DataFrame)
    """
    if len(dfs) != len(notas):
        raise ValueError("As listas 'dfs' e 'notas' devem ter o mesmo tamanho.")

    dfs_padrao = []

    for df, nota_coluna in zip(dfs, notas):
        # Verifica se colunas existem
        for col in ["INE", "NOME DA EQUIPE", nota_coluna]:
            if col not in df.columns:
                raise ValueError(f"Coluna '{col}' não encontrada no DataFrame.")

        # Cria novo DataFrame padronizado e renomeia a coluna de nota
        df_padrao = df[["INE", "NOME DA EQUIPE", nota_coluna]].copy()
        df_padrao = df_padrao.rename(columns={nota_coluna: nota_coluna.strip()})
        dfs_padrao.append(df_padrao)

    # Faz merge sequencial usando 'INE' e 'NOME DA EQUIPE' como chave
    from functools import reduce
    df_consolidado = reduce(
        lambda left, right: pd.merge(
            left, right, on=["INE", "NOME DA EQUIPE"], how="outer"
        ),
        dfs_padrao
    )

    return df_consolidado

def classificar_e_ordenar_desempate(
    df, 
    coluna_nota="NOTA FINAL", 
    coluna_saida="CLASSIFICACAO",
    desempate=["NOTA C2", "NOTA C3", "NOTA C7"]
):
    """
    Classifica notas em categorias e ordena o DataFrame pela nota final.
    Em caso de empate, desempata pelas colunas fornecidas na lista 'desempate'.
    
    Parâmetros:
    -----------
    df : pd.DataFrame
        DataFrame contendo as colunas de nota.
    coluna_nota : str
        Nome da coluna de nota final.
    coluna_saida : str
        Nome da coluna que será criada com a classificação.
    desempate : list of str
        Lista de colunas para desempate, em ordem de prioridade.

    Retorna:
    --------
    pd.DataFrame
        DataFrame com a coluna de classificação e ordenado pela nota final e desempate.
    """
    df = df.copy()
    
    # Garante que todas as colunas de notas sejam numéricas
    df[coluna_nota] = pd.to_numeric(df[coluna_nota], errors="coerce")
    for col in desempate:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    # Cria a coluna de classificação
    conditions = [
        df[coluna_nota] > 7.5,
        df[coluna_nota].between(5, 7.5),
        df[coluna_nota].between(2.6, 4.9),
        df[coluna_nota] <= 2.5
    ]
    
    choices = ["Ótimo", "Bom", "Suficiente", "Regular"]
    
    df[coluna_saida] = np.select(conditions, choices, default="Sem classificação")
    
    # Ordena: nota final decrescente e desempates decrescentes
    ordenar_colunas = [coluna_nota] + desempate
    df = df.sort_values(by=ordenar_colunas, ascending=False).reset_index(drop=True)
    
    return df

def remover_ultimas_linhas(df, n=1):
    """
    Remove as últimas n linhas de um DataFrame.
    """
    return df.iloc[:-n] if n > 0 else df.copy()

def _formatar_multilinhas(texto):
    if texto is None:
        return ""
    if isinstance(texto, (list, tuple)):
        return "\n".join(texto)
    return str(texto)


def exportar_excel(
    df,
    caminho,
    titulo=None,
    rodape=None,
    coluna_classificacao="CONCEITO",
    coluna_nota_final="NOTA FINAL",
    casas_decimais=2
):
    """
    Exporta DataFrame para Excel com:
    - Cabeçalho e rodapé flexíveis (1 a N linhas)
    - Texto colorido e em negrito conforme conceito
    - Formato numérico nas notas
    """

    # 1. Exportar DataFrame
    df.to_excel(caminho, index=False, engine="openpyxl")

    wb = load_workbook(caminho)
    ws = wb.active

    # 2. Cabeçalho e rodapé (flexíveis)
    titulo_txt = _formatar_multilinhas(titulo)
    rodape_txt = _formatar_multilinhas(rodape)

    if titulo_txt:
        ws.oddHeader.center.text = titulo_txt
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.size = 14

    if rodape_txt:
        ws.oddFooter.left.text = rodape_txt
        ws.oddFooter.left.font = "Arial"
        ws.oddFooter.left.size = 9

    # 3. Estilos de texto por conceito
    estilos = {
        "Regular": Font(color="ff2800", bold=True),
        "Suficiente": Font(color="FFA500", bold=True),
        "Bom": Font(color="04A514", bold=True),
        "Ótimo": Font(color="3129D0", bold=True)
    }

    headers = [cell.value for cell in ws[1]]

    idx_class = headers.index(coluna_classificacao) + 1
    idx_nota = headers.index(coluna_nota_final) + 1

    formato_num = f"0.{('0' * casas_decimais)}"

    # 4. Aplicar estilos
    for row in range(2, ws.max_row + 1):
        conceito = ws.cell(row=row, column=idx_class).value

        if conceito in estilos:
            fonte = estilos[conceito]

            ws.cell(row=row, column=idx_class).font = fonte

            cell_nota = ws.cell(row=row, column=idx_nota)
            cell_nota.font = fonte
            cell_nota.number_format = formato_num

    # 5. Largura padrão institucional
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 12

    wb.save(caminho)

    print(f"Relatório exportado com sucesso: {caminho}")




# def exportar_excel(df, caminho, titulo="", rodape="", engine="openpyxl"):
#     """
#     Exporta DataFrame para Excel com cabeçalho e rodapé visíveis apenas na impressão.
    
#     Parâmetros:
#     - df: DataFrame do pandas
#     - caminho: caminho do arquivo Excel
#     - titulo: texto do cabeçalho de impressão (centralizado)
#     - rodape: texto do rodapé de impressão (centralizado)
#     - engine: 'openpyxl' ou 'xlsxwriter' (atualmente openpyxl implementado)
#     """
#     # Salvar DataFrame em Excel
#     df.to_excel(caminho, index=False, engine="openpyxl")
    
#     # Abrir arquivo com openpyxl
#     wb = load_workbook(caminho)
#     ws = wb.active
    
#     # Cabeçalho de impressão
#     if titulo:
#         ws.oddHeader.center.text = titulo
#         ws.oddHeader.center.size = 14
#         ws.oddHeader.center.font = "Arial,Bold"
    
#     # Rodapé de impressão
#     if rodape:
#         ws.oddFooter.center.text = rodape
#         ws.oddFooter.center.size = 11
#         ws.oddFooter.center.font = "Arial"
    
#     wb.save(caminho)
#     print(f"Arquivo exportado com cabeçalho e rodapé de impressão em {caminho}")

# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill

# def exportar_excel(
#     df,
#     caminho,
#     titulo="",
#     rodape="",
#     coluna_classificacao="CLASSIFICACAO",
#     coluna_nota_final="NOTA FINAL",
#     engine="openpyxl"
# ):
#     """
#     Exporta DataFrame para Excel com:
#     - Cabeçalho e rodapé apenas para impressão
#     - Colunas de classificação e nota final coloridas e em negrito
#     """

#     # 1. Salvar DataFrame
#     df.to_excel(caminho, index=False, engine=engine)

#     # 2. Abrir com openpyxl
#     wb = load_workbook(caminho)
#     ws = wb.active

#     # 3. Cabeçalho de impressão
#     if titulo:
#         ws.oddHeader.center.text = titulo
#         ws.oddHeader.center.size = 14
#         ws.oddHeader.center.font = "Arial,Bold"

#     # 4. Rodapé de impressão
#     if rodape:
#         ws.oddFooter.center.text = rodape
#         ws.oddFooter.center.size = 11
#         ws.oddFooter.center.font = "Arial"

#     # 5. Mapeamento de cores
#     estilos = {
#         "Regular": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),     # vermelho claro
#         "Suficiente": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # laranja
#         "Bom": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),          # verde
#         "Ótimo": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")         # azul
#     }

#     fonte_negrito = Font(bold=True)

#     # 6. Descobrir índices das colunas
#     headers = [cell.value for cell in ws[1]]

#     idx_classificacao = headers.index(coluna_classificacao) + 1 if coluna_classificacao in headers else None
#     idx_nota = headers.index(coluna_nota_final) + 1 if coluna_nota_final in headers else None

#     # 7. Aplicar estilos linha a linha
#     for row in range(2, ws.max_row + 1):
#         if idx_classificacao:
#             valor_classificacao = ws.cell(row=row, column=idx_classificacao).value

#             if valor_classificacao in estilos:
#                 fill = estilos[valor_classificacao]

#                 # Classificação
#                 cell_class = ws.cell(row=row, column=idx_classificacao)
#                 cell_class.fill = fill
#                 cell_class.font = fonte_negrito

#                 # Nota final
#                 if idx_nota:
#                     cell_nota = ws.cell(row=row, column=idx_nota)
#                     cell_nota.fill = fill
#                     cell_nota.font = fonte_negrito

#     # 8. Salvar
#     wb.save(caminho)

#     print(f"Arquivo Excel exportado com estilos em: {caminho}")

# from openpyxl import load_workbook
# from openpyxl.styles import Font

# def exportar_excel(
#     df,
#     caminho,
#     titulo="",
#     rodape="",
#     coluna_classificacao="CONCEITO",
#     coluna_nota_final="NOTA FINAL",
#     casas_decimais=2
# ):
#     """
#     Exporta DataFrame para Excel com:
#     - Cabeçalho e rodapé apenas para impressão
#     - Cor aplicada SOMENTE ao texto
#     - Texto em negrito para classificação e nota final
#     - Formato numérico nas notas
#     """

#     # 1. Exporta DataFrame
#     df.to_excel(caminho, index=False, engine="openpyxl")

#     # 2. Abre com openpyxl
#     wb = load_workbook(caminho)
#     ws = wb.active

#     # 3. Cabeçalho de impressão
#     if titulo:
#         ws.oddHeader.center.text = titulo
#         ws.oddHeader.center.size = 14
#         ws.oddHeader.center.font = "Arial,Bold"

#     # 4. Rodapé de impressão
#     if rodape:
#         ws.oddFooter.center.text = rodape
#         ws.oddFooter.center.size = 10
#         ws.oddFooter.center.font = "Arial"

#     # 5. Cores SOMENTE no texto
#     cores_texto = {
#         "Regular": Font(color="E50743", bold=True),     # vermelho
#         "Suficiente": Font(color="FFA500", bold=True),  # laranja
#         "Bom": Font(color="09B474", bold=True),          # verde
#         "Ótimo": Font(color="3920F2", bold=True)         # azul
#     }

#     # 6. Identificar colunas
#     headers = [cell.value for cell in ws[1]]

#     idx_classificacao = headers.index(coluna_classificacao) + 1
#     idx_nota = headers.index(coluna_nota_final) + 1

#     formato_numero = f"0.{('0' * casas_decimais)}"

#     # 7. Aplicar estilos
#     for row in range(2, ws.max_row + 1):
#         conceito = ws.cell(row=row, column=idx_classificacao).value

#         if conceito in cores_texto:
#             fonte = cores_texto[conceito]

#             # Classificação
#             ws.cell(row=row, column=idx_classificacao).font = fonte

#             # Nota final
#             cell_nota = ws.cell(row=row, column=idx_nota)
#             cell_nota.font = fonte
#             cell_nota.number_format = formato_numero

#     # 8. Ajuste simples de largura (institucional)
#     for col in ws.columns:
#         ws.column_dimensions[col[0].column_letter].width = 11

#     wb.save(caminho)

#     print(f"Relatório exportado com sucesso em: {caminho}")





# import pandas as pd
# import numpy as np

# def classificar_e_ordenar(df, coluna_nota="NOTA_FINAL", coluna_saida="CLASSIFICACAO"):
#     """
#     Classifica notas em categorias e ordena o DataFrame pela nota final.

#     Parâmetros:
#     -----------
#     df : pd.DataFrame
#         DataFrame contendo a coluna de notas.
#     coluna_nota : str
#         Nome da coluna de notas a ser usada para classificação.
#     coluna_saida : str
#         Nome da coluna que será criada com a classificação.

#     Retorna:
#     --------
#     pd.DataFrame
#         DataFrame com a coluna de classificação e ordenado por nota final.
#     """
#     df = df.copy()
    
#     # Certifica que a coluna de nota é numérica
#     df[coluna_nota] = pd.to_numeric(df[coluna_nota], errors="coerce")
    
#     # Cria a coluna de classificação
#     conditions = [
#         df[coluna_nota] > 7.5,
#         df[coluna_nota].between(5, 7.5),
#         df[coluna_nota].between(2.6, 4.9),
#         df[coluna_nota] <= 2.5
#     ]
    
#     choices = ["Ótimo", "Bom", "Suficiente", "Regular"]
    
#     df[coluna_saida] = np.select(conditions, choices, default="Sem classificação")
    
#     # Ordena por nota final decrescente
#     df = df.sort_values(by=coluna_nota, ascending=False).reset_index(drop=True)
    
#     return df
